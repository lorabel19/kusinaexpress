# restaurant/views.py

# ============================================================================
# IMPORTS
# ============================================================================
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.db import connection
from django.db.models import Sum, Prefetch, Q
from django.utils import timezone
import re
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.hashers import make_password, check_password

from rest_framework import viewsets, status
from rest_framework.decorators import api_view
from rest_framework.response import Response

from decimal import Decimal
import json
from datetime import datetime

from .models import (
    MenuItems, Users, Orders, Cart, OrderItems, 
    ContactMessage, Feedback, Admin, Payments, Deliveries
)
from .serializers import (
    MenuItemsSerializer, CartSerializer, OrdersSerializer, FeedbackSerializer
)


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def get_logged_in_user(request):
    """Return currently logged-in regular user."""
    user_id = request.session.get('user_session_id')
    if not user_id:
        return None
    try:
        return Users.objects.get(user_id=user_id)
    except Users.DoesNotExist:
        # Clear only user session keys
        user_keys = [k for k in request.session.keys() if k.startswith('user_')]
        for key in user_keys:
            del request.session[key]
        return None

def get_logged_in_admin(request):
    """Return currently logged-in admin."""
    admin_id = request.session.get('admin_session_id')
    if not admin_id:
        return None
    try:
        return Admin.objects.get(admin_id=admin_id)
    except Admin.DoesNotExist:
        # Clear only admin session keys
        admin_keys = [k for k in request.session.keys() if k.startswith('admin_')]
        for key in admin_keys:
            del request.session[key]
        return None

def format_time(dt):
    """Convert datetime to string format."""
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else None


# ============================================================================
# AUTHENTICATION VIEWS
# ============================================================================

def login_view(request):
    """Handle user and admin login."""
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")
        role = request.POST.get("role")

        if role.lower() == 'admin':
            # Check Admin model
            try:
                admin = Admin.objects.get(email=email)
                
                if admin.check_password(password):
                    # Set admin session keys
                    request.session['admin_session_id'] = admin.admin_id
                    request.session['admin_role'] = 'admin'
                    request.session['admin_email'] = admin.email
                    request.session['admin_name'] = admin.name

                    messages.success(request, f"Welcome back, Admin {admin.name}!")
                    return redirect('restaurant:admin-dashboard')
                else:
                    messages.error(request, "Invalid email or password")
                    
            except Admin.DoesNotExist:
                messages.error(request, "No admin account found with this email")
                
        else:
            # Check Users model
            try:
                user = Users.objects.get(email=email)
                
                if check_password(password, user.password):
                    # Set user session keys
                    request.session['user_session_id'] = user.user_id
                    request.session['user_role'] = 'user'
                    request.session['user_email'] = user.email
                    request.session['user_first_name'] = user.first_name
                    request.session['user_last_name'] = user.last_name

                    messages.success(request, f"Welcome back, {user.first_name}!")
                    return redirect('restaurant:dashboard')
                else:
                    messages.error(request, "Invalid email or password")
                    
            except Users.DoesNotExist:
                messages.error(request, "Invalid email address")

    return render(request, 'restaurant/login.html')

def create_account(request):
    """Create new user account."""
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        
        # Validation
        errors = []
        
        # Check if email is valid
        if not email:
            errors.append('Email is required')
        elif not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            errors.append('Please enter a valid email address')
        elif Users.objects.filter(email=email).exists():
            errors.append('Email already exists')
        
        # Check password
        if not password:
            errors.append('Password is required')
        elif len(password) < 8:
            errors.append('Password must be at least 8 characters')
        elif not re.search(r'[A-Z]', password):
            errors.append('Password must contain at least one uppercase letter')
        elif not re.search(r'[a-z]', password):
            errors.append('Password must contain at least one lowercase letter')
        elif not re.search(r'[0-9]', password):
            errors.append('Password must contain at least one number')
        
        # Check password confirmation
        if password != confirm_password:
            errors.append('Passwords do not match')
        
        # Check names
        if not first_name:
            errors.append('First name is required')
        if not last_name:
            errors.append('Last name is required')
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return redirect('restaurant:login')
        
        try:
            # Create user with hashed password
            user = Users(
                email=email,
                password=make_password(password),
                first_name=first_name,
                last_name=last_name,
                role='user',
                date_joined=timezone.now(),
                is_active=1
            )
            user.save()
            
            messages.success(request, 'Account created successfully! You can now log in.')
            return redirect('restaurant:login')
            
        except Exception as e:
            messages.error(request, f'Error creating account: {str(e)}')
            return redirect('restaurant:login')
    
    return redirect('restaurant:login')

def user_logout_view(request):
    """Logout regular user."""
    user_keys = [k for k in request.session.keys() if k.startswith('user_')]
    for key in user_keys:
        del request.session[key]

    messages.success(request, "Logged out successfully!")
    return redirect('restaurant:login')

def admin_logout_view(request):
    """Logout admin."""
    admin_keys = [k for k in request.session.keys() if k.startswith('admin_')]
    for key in admin_keys:
        del request.session[key]

    messages.success(request, "Admin logged out successfully!")
    return redirect('restaurant:login')


# ============================================================================
# PUBLIC PAGES
# ============================================================================

def home(request):
    """Render home page."""
    user = get_logged_in_user(request)
    
    # Get featured menu items
    try:
        featured_items = MenuItems.objects.filter(
            is_available=1,
            category='meals'
        )[:3]
        
        featured_items_list = []
        for item in featured_items:
            featured_items_list.append({
                'name': item.name,
                'description': item.description or '',
                'price': float(item.price),
                'image_url': item.image_url or '',
                'is_available': item.is_available == 1,
                'item_id': item.item_id
            })
        
    except Exception as e:
        print(f"Error fetching featured items: {e}")
        featured_items_list = [
            {
                'name': 'Adobo',
                'description': 'Tender pork or chicken slowly cooked in soy sauce, vinegar, garlic, and bay leaves.',
                'price': 180,
                'image_url': 'https://images.deliveryhero.io/image/fd-ph/LH/t7o8-listing.jpg',
                'is_available': True,
                'item_id': 1
            },
            {
                'name': 'Sinigang',
                'description': 'A comforting sour soup made with tamarind, fresh vegetables, and your choice of pork.',
                'price': 200,
                'image_url': 'https://www.unileverfoodsolutions.com.ph/dam/global-ufs/mcos/SEA/calcmenu/recipes/PH-recipes/appetisers/sizzling-pork-sisig-manila/sizzling-pork-sisig-manila-main.jpg',
                'is_available': True,
                'item_id': 2
            }
        ]
    
    context = {
        'user': user,
        'featured_items': featured_items_list
    }
    
    return render(request, 'restaurant/index.html', context)

def about(request):
    """Render about page."""
    return render(request, 'restaurant/about.html')

def contact_page(request):
    """Render contact page."""
    return render(request, 'restaurant/contact.html')

def menu(request):
    """Render menu page with categories."""
    meals = MenuItems.objects.filter(category='meals')
    drinks = MenuItems.objects.filter(category='drinks')
    desserts = MenuItems.objects.filter(category='desserts')
    
    context = {
        'meals': meals,
        'drinks': drinks,
        'desserts': desserts
    }
    return render(request, 'restaurant/mainmenu.html', context)


# ============================================================================
# USER DASHBOARD & PROFILE
# ============================================================================

def dashboard(request):
    """Render dashboard for logged-in regular users."""
    user = get_logged_in_user(request)
    if not user:
        return redirect('restaurant:login')
    
    featured_items = MenuItems.objects.filter(is_available=1)[:6]
    
    context = {
        'user': user,
        'featured_items': featured_items
    }
    
    return render(request, 'restaurant/index_logged_in.html', context)

def profile_view(request):
    """Render user profile page."""
    user = get_logged_in_user(request)
    if not user:
        return redirect('restaurant:login')

    deliveries = Deliveries.objects.filter(
        order__user=user,
        delivered_at__isnull=False
    ).select_related('order').prefetch_related('order__items').order_by('-delivered_at')

    context = {
        'user': user,
        'deliveries': deliveries,
    }
    return render(request, 'restaurant/profile.html', context)



def update_profile(request):
    """Update user profile information."""
    # Use your custom function to get logged in user
    user = get_logged_in_user(request)
    if not user:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    try:
        # Get the user from database
        db_user = Users.objects.get(user_id=user.user_id)
        
        # Get form data
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip().lower()
        
        # Validate required fields
        if not first_name or not last_name or not email:
            return JsonResponse({'error': 'All fields are required'}, status=400)
        
        # Check if email is already taken by another user
        if email != db_user.email:
            if Users.objects.filter(email=email).exclude(user_id=db_user.user_id).exists():
                return JsonResponse({'error': 'Email is already registered'}, status=400)
        
        # Update user
        db_user.first_name = first_name
        db_user.last_name = last_name
        db_user.email = email
        
        # Save phone if provided (you need to add phone field to Users model first)
        # phone = request.POST.get('phone', '').strip()
        # if phone:
        #     db_user.phone = phone
        
        db_user.save()
        
        # Return success response with updated data
        return JsonResponse({
            'success': True,
            'message': 'Profile updated successfully',
            'first_name': db_user.first_name,
            'last_name': db_user.last_name,
            'email': db_user.email
        })
        
    except Users.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def change_password(request):
    """Change user password."""
    # Use your custom function to get logged in user
    user = get_logged_in_user(request)
    if not user:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    try:
        # Get the user from database
        db_user = Users.objects.get(user_id=user.user_id)
        
        # Get form data
        current_password = request.POST.get('current_password', '')
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')
        
        # Validate required fields
        if not current_password or not new_password or not confirm_password:
            return JsonResponse({'error': 'All fields are required'}, status=400)
        
        # Verify current password
        if not check_password(current_password, db_user.password):
            return JsonResponse({'error': 'Current password is incorrect'}, status=400)
        
        # Check if new password matches confirmation
        if new_password != confirm_password:
            return JsonResponse({'error': 'New passwords do not match'}, status=400)
        
        # Check password strength
        if len(new_password) < 8:
            return JsonResponse({'error': 'Password must be at least 8 characters long'}, status=400)
        
        # Check for password complexity (optional but recommended)
        import re
        if not re.search(r'[A-Z]', new_password):
            return JsonResponse({'error': 'Password must contain at least one uppercase letter'}, status=400)
        if not re.search(r'[a-z]', new_password):
            return JsonResponse({'error': 'Password must contain at least one lowercase letter'}, status=400)
        if not re.search(r'[0-9]', new_password):
            return JsonResponse({'error': 'Password must contain at least one number'}, status=400)
        
        # Update password
        db_user.password = make_password(new_password)
        db_user.save()
        
        # Return success response
        return JsonResponse({
            'success': True,
            'message': 'Password changed successfully'
        })
        
    except Users.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)




# ============================================================================
# MENU ITEMS API
# ============================================================================

class MenuItemsViewSet(viewsets.ModelViewSet):
    """ViewSet for managing menu items via REST API."""
    queryset = MenuItems.objects.all()
    serializer_class = MenuItemsSerializer

def menu_page(request):
    """Render menu page."""
    user = get_logged_in_user(request)
    items = MenuItems.objects.filter(is_available=1)
    return render(request, 'restaurant/menu.html', {'user': user, 'menu_items': items})


# ============================================================================
# CART VIEWS & APIs
# ============================================================================

def cart_page(request):
    """Render cart page."""
    user = get_logged_in_user(request)
    if not user:
        return redirect('restaurant:login')

    cart_items = Cart.objects.filter(user=user)
    total = sum(item.subtotal for item in cart_items)
    return render(request, 'restaurant/cart.html', {'user': user, 'cart_items': cart_items, 'total': total})

@api_view(['GET'])
def cart_api(request):
    """Return current user's cart items."""
    user = get_logged_in_user(request)
    if not user:
        return Response({"detail": "Not logged in"}, status=status.HTTP_401_UNAUTHORIZED)

    items = Cart.objects.filter(user=user)
    serializer = CartSerializer(items, many=True)
    return Response(serializer.data)

@api_view(['POST'])
def add_to_cart_api(request):
    """Add a menu item to the cart."""
    user = get_logged_in_user(request)
    if not user:
        return Response({"detail": "Not logged in"}, status=status.HTTP_401_UNAUTHORIZED)

    item_id = request.data.get('item_id')
    quantity = int(request.data.get('quantity', 1))

    if not item_id:
        return Response({"detail": "Item ID is required"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        menu_item = MenuItems.objects.get(item_id=item_id)
        cart_item, created = Cart.objects.get_or_create(user=user, item=menu_item)

        # Update quantity
        if created:
            cart_item.quantity = quantity
        else:
            cart_item.quantity = (cart_item.quantity or 0) + quantity

        # Delete cart item if quantity <= 0
        if cart_item.quantity <= 0:
            cart_item.delete()
            return Response({"detail": f"'{menu_item.name}' removed from cart."}, status=status.HTTP_200_OK)

        # Compute subtotal
        cart_item.subtotal = Decimal(menu_item.price) * Decimal(cart_item.quantity)
        cart_item.save()

        return Response({
            "detail": f"'{menu_item.name}' added/updated in cart.",
            "cart_id": cart_item.cart_id,
            "item_id": menu_item.item_id,
            "name": menu_item.name,
            "price": str(menu_item.price),
            "quantity": cart_item.quantity,
            "subtotal": str(cart_item.subtotal),
            "image_url": menu_item.image_url
        }, status=status.HTTP_200_OK)

    except MenuItems.DoesNotExist:
        return Response({"detail": "Menu item not found"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"detail": f"Error: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['DELETE'])
def remove_from_cart_api(request, cart_id):
    """Remove an item from cart."""
    user = get_logged_in_user(request)
    if not user:
        return Response({"detail": "Not logged in"}, status=status.HTTP_401_UNAUTHORIZED)

    try:
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM cart WHERE cart_id = %s AND user_id = %s", [cart_id, user.user_id])
        return Response({"detail": "Item removed successfully"}, status=status.HTTP_200_OK)

    except Exception as e:
        return Response({"detail": f"Error removing item: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================================
# ORDER PLACEMENT & TRACKING
# ============================================================================

@csrf_exempt
def place_order_api(request):
    """Place a new order from cart items."""
    if request.method != "POST":
        return JsonResponse({"detail": "Method not allowed"}, status=405)

    user = get_logged_in_user(request)
    if not user:
        return JsonResponse({"detail": "Not logged in"}, status=401)

    cart_items = Cart.objects.filter(user_id=user.user_id)
    if not cart_items.exists():
        return JsonResponse({"detail": "No items in cart"}, status=400)

    try:
        data = json.loads(request.body)

        # Calculate total
        subtotal = sum([ci.subtotal for ci in cart_items])
        shipping_fee = Decimal('40.00')
        total_amount = subtotal + shipping_fee

        now = timezone.localtime()

        # Create order
        order = Orders.objects.create(
            user_id=user.user_id,
            total_amount=total_amount,
            order_date=now
        )

        # Save order items
        for ci in cart_items:
            OrderItems.objects.create(
                order_id=order.order_id,
                item_id=ci.item.item_id,
                quantity=ci.quantity,
                subtotal=ci.subtotal
            )

        # Save payment
        Payments.objects.create(
            order=order,
            payment_method=data.get("payment_method", ""),
            payment_date=now
        )

        # Save delivery
        Deliveries.objects.create(
            order=order,
            delivery_address=data.get("address", ""),
            contact_number=data.get("contact", ""),
            delivery_option=data.get("delivery_option", ""),
            notes=data.get("notes", ""),
            status="pending",
            confirmed_at=None,
            preparing_at=None,
            out_for_delivery_at=None,
            delivered_at=None
        )

        # Clear cart
        cart_items.delete()

        return JsonResponse({
            "detail": "Order placed successfully",
            "order_id": order.order_id
        }, status=201)

    except Exception as e:
        return JsonResponse({"detail": f"Error placing order: {str(e)}"}, status=500)

def order_view(request):
    """Render order tracking page."""
    user = get_logged_in_user(request)
    if not user:
        return render(
            request,
            "restaurant/order.html",
            {"order": None, "order_id": None},
        )

    order = Orders.objects.filter(user=user).order_by("-order_date").first()

    return render(
        request,
        "restaurant/order.html",
        {
            "order": order,
            "order_id": order.order_id if order else None,
        },
    )

@api_view(["GET"])
def track_order_api(request, order_id):
    """Return order status and tracking information."""
    order = get_object_or_404(Orders, pk=order_id)
    delivery = Deliveries.objects.filter(order=order).first()

    # Demo map points
    START_COORDS = [14.6760, 121.0437]  # Quezon City
    END_COORDS = [14.5176, 121.0509]    # Paranaque

    steps = [
        "Order Confirmed",
        "Preparing Order",
        "Out for Delivery",
        "Delivered",
    ]

    def format_time(dt):
        return timezone.localtime(dt).strftime("%I:%M %p") if dt else None

    timestamps = {}
    status = "Pending"
    if delivery:
        timestamps = {
            "Order Confirmed": format_time(delivery.confirmed_at),
            "Preparing Order": format_time(delivery.preparing_at),
            "Out for Delivery": format_time(delivery.out_for_delivery_at),
            "Delivered": format_time(delivery.delivered_at),
        }
        status = delivery.status or "Pending"

    return Response(
        {
            "order_id": order.order_id,
            "status": status,
            "order_date": format_time(order.order_date),
            "total_amount": float(order.total_amount),
            "steps": steps,
            "timestamps": timestamps,
            "start_lat": START_COORDS[0],
            "start_lng": START_COORDS[1],
            "end_lat": END_COORDS[0],
            "end_lng": END_COORDS[1],
        }
    )

@csrf_exempt
def mark_order_seen(request, order_id):
    """Mark an order as seen by user."""
    if request.method != "POST":
        return JsonResponse({"detail": "Method not allowed"}, status=405)

    try:
        order = Orders.objects.get(pk=order_id)
        delivery = Deliveries.objects.filter(order=order).first()
        if delivery:
            delivery.timestamps = {}
            delivery.save()
        return JsonResponse({"detail": "Order marked as seen"})
    except Orders.DoesNotExist:
        return JsonResponse({"detail": "Order not found"}, status=404)


# ============================================================================
# CONTACT & FEEDBACK
# ============================================================================


@api_view(['POST'])
def submit_feedback(request):
    """Submit user feedback."""
    user_id = request.session.get('user_session_id')
    if not user_id:
        return Response({'detail': 'You must be logged in to submit feedback.'}, status=status.HTTP_403_FORBIDDEN)

    user = Users.objects.get(pk=user_id)
    serializer = FeedbackSerializer(data=request.data, context={'request': request})
    if serializer.is_valid():
        serializer.save(user=user, date_submitted=timezone.now())
        return Response({'message': 'Feedback submitted successfully!'}, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ============================================================================
# ADMIN DASHBOARD
# ============================================================================

def admin_dashboard(request):
    """Render admin dashboard with statistics."""
    admin = get_logged_in_admin(request)
    if not admin:
        return redirect('restaurant:login')

    # Summary counts
    total_orders = Orders.objects.count()
    pending_orders = Deliveries.objects.filter(status='pending').count()
    total_menu = MenuItems.objects.count()
    total_users = Users.objects.count()

    # Last 10 orders
    orders = Orders.objects.prefetch_related(
        Prefetch('deliveries_set', queryset=Deliveries.objects.all()),
        Prefetch('items', queryset=OrderItems.objects.select_related('item'))
    ).order_by('-order_date')[:10]

    # Best Sellers (top 5)
    best_sellers = (
        OrderItems.objects
        .values('item', 'item__name', 'item__image_url')
        .annotate(total_ordered=Sum('quantity'))
        .order_by('-total_ordered')[:5]
    )

    # Low Sellers (bottom 5)
    low_sellers = (
        OrderItems.objects
        .values('item', 'item__name', 'item__image_url')
        .annotate(total_ordered=Sum('quantity'))
        .order_by('total_ordered')[:5]
    )

    context = {
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'total_menu': total_menu,
        'total_users': total_users,
        'orders': orders,
        'best_sellers': best_sellers,
        'low_sellers': low_sellers,
        'admin': admin,
    }

    return render(request, 'restaurant/admin_dashboard.html', context)


# ============================================================================
# ADMIN MENU MANAGEMENT
# ============================================================================

def admin_menu(request):
    """Display all menu items for admin."""
    admin = get_logged_in_admin(request)
    if not admin:
        return redirect('restaurant:login')
    
    menu_items = MenuItems.objects.all()
    return render(request, 'restaurant/admin_menu.html', {'menu_items': menu_items, 'admin': admin})

def admin_add_menu(request):
    """Add new menu item."""
    admin = get_logged_in_admin(request)
    if not admin:
        return redirect('restaurant:login')
    
    if request.method == "POST":
        name = request.POST.get("name")
        description = request.POST.get("description")
        price = request.POST.get("price")
        category = request.POST.get("category")
        image_url = request.POST.get("image_url")
        is_available = int(request.POST.get("is_available", 1))

        MenuItems.objects.create(
            name=name,
            description=description,
            price=price,
            category=category,
            image_url=image_url,
            is_available=is_available
        )
        messages.success(request, f"{name} added successfully!")
        return redirect('restaurant:admin-menu')

    return render(request, 'restaurant/admin_menu_add.html', {'admin': admin})

def admin_edit_menu(request, item_id):
    """Edit existing menu item."""
    admin = get_logged_in_admin(request)
    if not admin:
        return redirect('restaurant:login')
    
    item = get_object_or_404(MenuItems, item_id=item_id)

    if request.method == "POST":
        item.name = request.POST.get("name")
        item.description = request.POST.get("description")
        item.price = request.POST.get("price")
        item.category = request.POST.get("category")
       
        item.is_available = int(request.POST.get("is_available", 1))
        item.save()
        messages.success(request, f"{item.name} updated successfully!")
        return redirect('restaurant:admin-menu')

    return render(request, 'restaurant/admin_menu_edit.html', {'item': item, 'admin': admin})

from django.db import transaction
from django.db import connection

def admin_delete_menu(request, item_id):
    """Delete a menu item."""
    admin = get_logged_in_admin(request)
    if not admin:
        return redirect('restaurant:login')
    
    try:
        with transaction.atomic():
            from .models import MenuItems
            
            # Get the item name first
            item = MenuItems.objects.get(item_id=item_id)
            item_name = item.name
            
            # Use raw SQL to bypass Django's model constraints
            with connection.cursor() as cursor:
                # Delete order items first
                cursor.execute("DELETE FROM OrderItems WHERE item_id = %s", [item_id])
                
                # Delete menu item
                cursor.execute("DELETE FROM menu_items WHERE item_id = %s", [item_id])
            
            messages.success(request, f"{item_name} deleted successfully!")
            
    except MenuItems.DoesNotExist:
        messages.error(request, "Item not found!")
    except Exception as e:
        messages.error(request, f"Error deleting item: {str(e)}")
    
    return redirect('restaurant:admin-menu')


# ============================================================================
# ADMIN ORDER MANAGEMENT
# ============================================================================

from datetime import timedelta

def admin_orders(request):
    """Display orders filtered by status for admin."""
    admin = get_logged_in_admin(request)
    if not admin:
        return redirect('restaurant:login')
    
    status_filter = request.GET.get('status', 'pending')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Base queryset with prefetch
    orders = Orders.objects.all().prefetch_related(
        Prefetch('items', queryset=OrderItems.objects.select_related('item')),
        Prefetch('deliveries_set', queryset=Deliveries.objects.all()),
        Prefetch('payments_set', queryset=Payments.objects.all())
    )

    # Apply status filter
    if status_filter != 'all':
        if status_filter == 'delivered':
            orders = orders.filter(deliveries__delivered_at__isnull=False)
        elif status_filter == 'out_for_delivery':
            orders = orders.filter(deliveries__status='out_for_delivery')
        elif status_filter == 'preparing':
            orders = orders.filter(deliveries__status='preparing')
        elif status_filter == 'confirmed':
            orders = orders.filter(deliveries__status='confirmed')
        elif status_filter == 'pending':
            orders = orders.filter(
                Q(deliveries__isnull=True) | 
                Q(deliveries__status='pending')
            )
        else:
            orders = orders.filter(deliveries__status=status_filter)

    # Apply date filtering for delivered orders
    if status_filter == 'delivered' and (start_date or end_date):
        if start_date:
            try:
                start_date_obj = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
                orders = orders.filter(deliveries__delivered_at__gte=start_date_obj)
            except ValueError:
                pass
        
        if end_date:
            try:
                end_date_obj = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d'))
                end_date_obj = end_date_obj + timedelta(days=1)
                orders = orders.filter(deliveries__delivered_at__lt=end_date_obj)
            except ValueError:
                pass

    # Order by order_date
    orders = orders.order_by('-order_date')
    orders = [order for order in orders if order.items.exists()]

    return render(
        request,
        "restaurant/admin_orders.html",
        {
            'orders': orders,
            'status_filter': status_filter,
            'admin': admin,
            'start_date': start_date,
            'end_date': end_date
        }
    )

def admin_update_orders(request, order_id):
    """Update order status and sync Deliveries timestamps."""
    admin = get_logged_in_admin(request)
    if not admin:
        return redirect('restaurant:login')
    
    order = get_object_or_404(Orders, order_id=order_id)
    delivery = Deliveries.objects.filter(order=order).first()

    if request.method == "POST" and delivery:
        new_status = request.POST.get("status")
        now = timezone.now()
        
        if new_status == "preparing":
            delivery.status = "preparing"
            delivery.preparing_at = now
            if not delivery.confirmed_at:
                delivery.confirmed_at = now

        elif new_status == "out_for_delivery":
            delivery.status = "out_for_delivery"
            delivery.out_for_delivery_at = now
            if not delivery.confirmed_at:
                delivery.confirmed_at = now
            if not delivery.preparing_at:
                delivery.preparing_at = now

        elif new_status == "delivered":
            delivery.status = "delivered"
            delivery.delivered_at = now
            if not delivery.confirmed_at:
                delivery.confirmed_at = now
            if not delivery.preparing_at:
                delivery.preparing_at = now
            if not delivery.out_for_delivery_at:
                delivery.out_for_delivery_at = now

        delivery.save()
        
        return redirect(f'/admin-orders/?status={new_status}')

    return redirect('/admin-orders/')

def confirm_order(request, order_id):
    """Confirm order."""
    admin = get_logged_in_admin(request)
    if not admin:
        return redirect('restaurant:login')
    
    order = get_object_or_404(Orders, order_id=order_id)
    
    if request.method == "POST":
        now = timezone.now()
        
        delivery, created = Deliveries.objects.get_or_create(
            order=order,
            defaults={
                'status': 'confirmed',
                'confirmed_at': now
            }
        )
        
        if not created:
            delivery.status = 'confirmed'
            delivery.confirmed_at = now
            delivery.save()

    return redirect('/admin-orders/?status=confirmed')


# ============================================================================
# ADMIN FEEDBACK MANAGEMENT
# ============================================================================

def admin_feedback(request):
    """Display all user feedback for admin."""
    admin = get_logged_in_admin(request)
    if not admin:
        return redirect('restaurant:login')
    
    feedback_list = Feedback.objects.all().order_by('-date_submitted')
    context = {
        'feedback_list': feedback_list,
        'admin': admin
    }
    return render(request, 'restaurant/admin_feedback.html', context)


# ============================================================================
# ADMIN USER MANAGEMENT
# ============================================================================

def manage_users(request):
    """
    Manage users and admins.
    - Users tab: All Users table entries
    - Admins tab: All Admin table entries
    """
    admin = get_logged_in_admin(request)
    if not admin:
        return redirect('restaurant:login')
    
    users = Users.objects.all().order_by('-date_joined')
    admins = Admin.objects.all().order_by('-created_at')
    
    success_message = request.session.pop('success_message', None)
    error_message = request.session.pop('error_message', None)
    
    context = {
        'users': users,
        'admins': admins,
        'admin': admin,
        'success_message': success_message,
        'error_message': error_message,
    }
    
    return render(request, 'restaurant/manage_users.html', context)

def add_admin(request):
    """Add new admin account."""
    admin_user = get_logged_in_admin(request)
    if not admin_user:
        return redirect('restaurant:login')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validate passwords match
        if password != confirm_password:
            request.session['error_message'] = 'Passwords do not match!'
            return redirect('restaurant:admin-users')
        
        # Check if email already exists in Admin table
        if Admin.objects.filter(email=email).exists():
            request.session['error_message'] = 'Email already exists in admin accounts!'
            return redirect('restaurant:admin-users')
        
        # Also check if email exists in Users table
        if Users.objects.filter(email=email).exists():
            request.session['error_message'] = 'Email already exists in users!'
            return redirect('restaurant:admin-users')
        
        # Create new admin
        try:
            admin = Admin(
                name=name,
                email=email,
                password=make_password(password)
            )
            admin.save()
            request.session['success_message'] = 'Admin added successfully!'
            return redirect('restaurant:admin-users')
        except Exception as e:
            print(f"Error creating admin: {e}")
            request.session['error_message'] = 'Server error. Please try again.'
            return redirect('restaurant:admin-users')
    
    return redirect('restaurant:admin-users')

def edit_admin(request):
    """Edit existing admin account."""
    admin_user = get_logged_in_admin(request)
    if not admin_user:
        return redirect('restaurant:login')
    
    if request.method == 'POST':
        admin_id = request.POST.get('admin_id')
        name = request.POST.get('name')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        
        try:
            admin = Admin.objects.get(admin_id=admin_id)
            
            # Update basic info
            admin.name = name
            admin.email = email
            
            # Update password if provided
            if password and confirm_password:
                if password != confirm_password:
                    request.session['error_message'] = 'Passwords do not match!'
                    return redirect('restaurant:admin-users')
                if len(password) < 6:
                    request.session['error_message'] = 'Password must be at least 6 characters long!'
                    return redirect('restaurant:admin-users')
                admin.password = make_password(password)
            
            # Check if email already exists
            if Admin.objects.filter(email=email).exclude(admin_id=admin_id).exists():
                request.session['error_message'] = 'Email already exists!'
                return redirect('restaurant:admin-users')
            
            admin.save()
            request.session['success_message'] = 'Admin updated successfully!'
            return redirect('restaurant:admin-users')
            
        except Admin.DoesNotExist:
            request.session['error_message'] = 'Admin not found!'
            return redirect('restaurant:admin-users')
        except Exception as e:
            print(f"Error updating admin: {e}")
            request.session['error_message'] = 'Server error. Please try again.'
            return redirect('restaurant:admin-users')
    
    return redirect('restaurant:admin-users')

def delete_admin(request):
    """Delete admin account."""
    admin_user = get_logged_in_admin(request)
    if not admin_user:
        return redirect('restaurant:login')
    
    if request.method == 'POST':
        admin_id = request.POST.get('admin_id')
        
        try:
            admin = Admin.objects.get(admin_id=admin_id)
            admin_name = admin.name
            admin.delete()
            request.session['success_message'] = f'Admin "{admin_name}" deleted successfully!'
            return redirect('restaurant:admin-users')
            
        except Admin.DoesNotExist:
            request.session['error_message'] = 'Admin not found!'
            return redirect('restaurant:admin-users')
        except Exception as e:
            print(f"Error deleting admin: {e}")
            request.session['error_message'] = 'Server error. Please try again.'
            return redirect('restaurant:admin-users')
    
    return redirect('restaurant:admin-users')


# ============================================================================
# ADMIN SETTINGS
# ============================================================================

def admin_settings(request):
    """Admin profile settings page."""
    admin = get_logged_in_admin(request)
    if not admin:
        messages.error(request, 'Please login to access settings')
        return redirect('restaurant:login')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validate required fields
        if not name or not email:
            messages.error(request, 'Name and email are required!')
            return redirect('restaurant:admin-settings')
        
        # Validate email format
        if '@' not in email:
            messages.error(request, 'Please enter a valid email address!')
            return redirect('restaurant:admin-settings')
        
        # Validate email uniqueness if changed
        if email != admin.email:
            if Admin.objects.filter(email=email).exclude(admin_id=admin.admin_id).exists():
                messages.error(request, 'Email already exists!')
                return redirect('restaurant:admin-settings')
        
        # Update basic info
        admin.name = name
        admin.email = email
        
        # Handle password change
        password_changed = False
        if current_password or new_password or confirm_password:
            # If any password field is filled, all must be filled
            if not (current_password and new_password and confirm_password):
                messages.error(request, 'Please fill all password fields to change password!')
                return redirect('restaurant:admin-settings')
            
            if admin.check_password(current_password):
                if new_password == confirm_password:
                    # Validate password strength
                    if len(new_password) < 8:
                        messages.error(request, 'Password must be at least 8 characters long!')
                        return redirect('restaurant:admin-settings')
                    
                    admin.password = make_password(new_password)
                    password_changed = True
                    messages.success(request, 'Password updated successfully!')
                else:
                    messages.error(request, 'New passwords do not match!')
                    return redirect('restaurant:admin-settings')
            else:
                messages.error(request, 'Current password is incorrect!')
                return redirect('restaurant:admin-settings')
        
        try:
            admin.save()
            messages.success(request, 'Profile updated successfully!')
            
            # Update session variables
            request.session['admin_name'] = admin.name
            request.session['admin_email'] = admin.email
            
            return redirect('restaurant:admin-settings')
        except Exception as e:
            messages.error(request, f'Error updating profile: {str(e)}')
    
    return render(request, 'restaurant/admin_settings.html', {'admin': admin})

from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json

@method_decorator(csrf_exempt, name='dispatch')
class FeedbackView(View):
    def post(self, request):
        user_id = request.session.get('user_session_id')
        if not user_id:
            return JsonResponse({'detail': 'You must be logged in to submit feedback.'}, status=403)
        
        try:
            user = Users.objects.get(pk=user_id)
        except Users.DoesNotExist:
            return JsonResponse({'detail': 'User not found'}, status=404)
        
        # Parse request data
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                data = {}
        else:
            data = {
                'message': request.POST.get('message'),
                'rating': request.POST.get('rating')
            }
        
        # Validate required field
        if not data.get('message'):
            return JsonResponse({'message': ['This field is required.']}, status=400)
        
        # Create feedback
        feedback = Feedback.objects.create(
            user=user,
            message=data['message'],
            rating=int(data['rating']) if data.get('rating') and data['rating'].isdigit() else None,
            date_submitted=timezone.now()
        )
        
        return JsonResponse({
            'message': 'Feedback submitted successfully!',
            'feedback_id': feedback.feedback_id
        }, status=201)
    
    from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone

@csrf_exempt
def submit_feedback_form(request):
    """Traditional form submission for feedback (not API)."""
    if request.method != 'POST':
        return JsonResponse({'detail': 'Method not allowed'}, status=405)
    
    user_id = request.session.get('user_session_id')
    if not user_id:
        return JsonResponse({'detail': 'You must be logged in to submit feedback.'}, status=403)

    try:
        user = Users.objects.get(pk=user_id)
    except Users.DoesNotExist:
        return JsonResponse({'detail': 'User not found'}, status=404)
    
    # Get form data
    message = request.POST.get('message', '').strip()
    rating = request.POST.get('rating', '').strip()
    
    # Validate
    if not message:
        return JsonResponse({'error': 'Message is required'}, status=400)
    
    # Validate rating
    rating_int = None
    if rating:
        try:
            rating_int = int(rating)
            if rating_int < 1 or rating_int > 5:
                return JsonResponse({'error': 'Rating must be between 1 and 5'}, status=400)
        except ValueError:
            return JsonResponse({'error': 'Invalid rating value'}, status=400)
    
    # Create feedback
    try:
        Feedback.objects.create(
            user=user,
            message=message,
            rating=rating_int,
            date_submitted=timezone.now()
        )
        return JsonResponse({'message': 'Feedback submitted successfully!'}, status=201)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    

# restaurant/views.py
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from datetime import datetime
import csv
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import openpyxl
from .models import Orders, Deliveries

def export_orders(request):
    """Handle order export requests"""
    try:
        report_type = request.GET.get('report_type', 'detailed')
        format_type = request.GET.get('format', 'pdf')
        status = request.GET.get('status', 'delivered')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        print(f"Export Parameters: status={status}, start_date={start_date}, end_date={end_date}, format={format_type}, report_type={report_type}")
        
        # Start with all orders
        orders = Orders.objects.all()
        
        # Apply status filter based on delivery status
        if status == 'delivered':
            orders = orders.filter(deliveries__delivered_at__isnull=False)
            print(f"Filtering delivered orders. Found: {orders.count()}")
        elif status == 'out_for_delivery':
            orders = orders.filter(deliveries__out_for_delivery_at__isnull=False, 
                                   deliveries__delivered_at__isnull=True)
        elif status == 'preparing':
            orders = orders.filter(deliveries__preparing_at__isnull=False,
                                   deliveries__out_for_delivery_at__isnull=True)
        elif status == 'confirmed':
            orders = orders.filter(deliveries__confirmed_at__isnull=False,
                                   deliveries__preparing_at__isnull=True)
        elif status == 'pending':
            orders = orders.filter(deliveries__confirmed_at__isnull=True)
        else:
            # If no status specified, show all
            pass
        
        # Apply date filter if provided - FIXED
        if start_date and end_date:
            # Convert string dates to datetime objects
            from datetime import datetime
            try:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                # Add time to end date to include the entire day
                from datetime import timedelta
                end_datetime = end_datetime + timedelta(days=1) - timedelta(seconds=1)
                
                print(f"Date filter range: {start_datetime} to {end_datetime}")
                
                if status == 'delivered':
                    # For delivered orders, filter by delivered_at date
                    orders = orders.filter(deliveries__delivered_at__range=[start_datetime, end_datetime])
                    print(f"Filtered by delivered_at. Found: {orders.count()}")
                else:
                    # For other statuses, filter by order_date
                    orders = orders.filter(order_date__range=[start_datetime, end_datetime])
                    print(f"Filtered by order_date. Found: {orders.count()}")
                    
            except Exception as e:
                print(f"Date parsing error: {str(e)}")
                # Fallback to string comparison if datetime parsing fails
                orders = orders.filter(order_date__date__range=[start_date, end_date])
                print(f"Fallback date filter. Found: {orders.count()}")
        
        # Sort by order date (newest first)
        orders = orders.order_by('-order_date')
        
        print(f"Total orders for export: {orders.count()}")
        
        # Debug: Print sample orders
        sample_count = min(3, orders.count())
        for i, order in enumerate(orders[:sample_count]):
            delivery = order.deliveries_set.first() if hasattr(order, 'deliveries_set') and order.deliveries_set.exists() else None
            print(f"Sample Order {i+1}: #{order.order_id}, Date: {order.order_date}, "
                  f"Delivery Date: {delivery.delivered_at if delivery else 'None'}")
        
        if format_type == 'pdf':
            if report_type == 'receipts':
                print("Generating individual receipts PDF...")
                return generate_individual_receipts_pdf(orders)
            else:
                print("Generating detailed PDF report...")
                return generate_detailed_pdf_report(orders, report_type, status)
        elif format_type == 'excel':
            print("Generating Excel report...")
            return generate_detailed_excel_report(orders, report_type, status)
        elif format_type == 'csv':
            print("Generating CSV report...")
            return generate_detailed_csv_report(orders, report_type, status)
        else:
            print("Default: Generating detailed PDF report...")
            return generate_detailed_pdf_report(orders, report_type, status)
            
    except Exception as e:
        import traceback
        print(f"Export error: {str(e)}")
        print(traceback.format_exc())
        return HttpResponse(f"Error generating report: {str(e)}", status=500)

def generate_detailed_pdf_report(orders, report_type, status):
    """Generate detailed PDF report with compact layout"""
    try:
        buffer = BytesIO()
        
        # Create PDF document with smaller margins
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=letter,
            leftMargin=0.4*inch,
            rightMargin=0.4*inch,
            topMargin=0.3*inch,
            bottomMargin=0.3*inch
        )
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Compact Title style
        title_style = ParagraphStyle(
            'CompactTitle',
            parent=styles['Heading1'],
            fontSize=12,
            textColor=colors.HexColor('#D62828'),
            alignment=TA_CENTER,
            spaceAfter=4
        )
        
        # Compact Subtitle style
        subtitle_style = ParagraphStyle(
            'CompactSubtitle',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.gray,
            alignment=TA_CENTER,
            spaceAfter=6
        )
        
        # Compact Order Header style
        order_header_style = ParagraphStyle(
            'CompactOrderHeader',
            parent=styles['Heading2'],
            fontSize=9,
            textColor=colors.HexColor('#D62828'),
            spaceAfter=2,
            spaceBefore=4
        )
        
        # Compact Label style
        label_style = ParagraphStyle(
            'CompactLabel',
            parent=styles['Normal'],
            fontSize=6,
            textColor=colors.HexColor('#6C757D'),
            spaceAfter=0
        )
        
        # Compact Value style
        value_style = ParagraphStyle(
            'CompactValue',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.black,
            spaceAfter=1
        )
        
        # Compact Item style
        item_style = ParagraphStyle(
            'CompactItem',
            parent=styles['Normal'],
            fontSize=6,
            leftIndent=8,
            spaceAfter=0
        )
        
        # Title
        status_text = status.upper().replace('_', ' ') if status else 'ALL'
        title = Paragraph(f"KUSINAEXPRESS - {status_text} ORDERS", title_style)
        elements.append(title)
        
        # Subtitle
        subtitle = Paragraph(f"Generated: {datetime.now().strftime('%m/%d/%Y %I:%M%p')}", subtitle_style)
        elements.append(subtitle)
        
        # Order count and total
        total_amount = sum(order.total_amount for order in orders)
        count_text = Paragraph(f"Orders: {orders.count()} | Total: ₱{total_amount:,.2f}", subtitle_style)
        elements.append(count_text)
        
        # Add minimal space
        elements.append(Spacer(1, 8))
        
        if orders.count() == 0:
            no_orders = Paragraph("No orders found.", 
                                 ParagraphStyle('NoOrders', parent=styles['Normal'], 
                                               fontSize=8, alignment=TA_CENTER, textColor=colors.gray))
            elements.append(no_orders)
            doc.build(elements)
            buffer.seek(0)
            filename = f"orders_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
        # Process each order
        for i, order in enumerate(orders):
            try:
                delivery = order.deliveries_set.first() if hasattr(order, 'deliveries_set') and order.deliveries_set.exists() else None
                
                # Order Header
                order_header = Paragraph(f"<b>#{order.order_id}</b>", order_header_style)
                elements.append(order_header)
                
                # Create a compact details table
                details_data = []
                
                # Status
                if delivery and hasattr(delivery, 'delivered_at') and delivery.delivered_at:
                    status_text = "DELIVERED"
                    status_color = colors.green
                elif delivery and hasattr(delivery, 'out_for_delivery_at') and delivery.out_for_delivery_at:
                    status_text = "OUT FOR DELIVERY"
                    status_color = colors.blue
                elif delivery and hasattr(delivery, 'preparing_at') and delivery.preparing_at:
                    status_text = "PREPARING"
                    status_color = colors.purple
                elif delivery and hasattr(delivery, 'confirmed_at') and delivery.confirmed_at:
                    status_text = "CONFIRMED"
                    status_color = colors.blue
                else:
                    status_text = "PENDING"
                    status_color = colors.orange
                
                details_data.append(['Status:', f"<font color='{status_color}'><b>{status_text}</b></font>"])
                
                # Customer
                customer_name = f"{order.user.first_name or ''} {order.user.last_name or ''}".strip()
                if customer_name:
                    details_data.append(['Customer:', customer_name[:20] + '..' if len(customer_name) > 20 else customer_name])
                
                # Order Date (compact format)
                order_date = order.order_date.strftime('%m/%d %H:%M') if hasattr(order, 'order_date') else "N/A"
                details_data.append(['Date:', order_date])
                
                # Delivery Address (truncated)
                if delivery and hasattr(delivery, 'delivery_address') and delivery.delivery_address:
                    address = delivery.delivery_address
                    if len(address) > 25:
                        address = address[:22] + "..."
                    details_data.append(['Address:', address])
                
                # Contact Number
                if delivery and hasattr(delivery, 'contact_number') and delivery.contact_number:
                    details_data.append(['Contact:', delivery.contact_number])
                
                # Payment Method
                payment_method = "COD"
                if hasattr(order, 'payments_set') and order.payments_set.exists():
                    payment = order.payments_set.first()
                    if payment and hasattr(payment, 'payment_method'):
                        payment_method = payment.payment_method or "COD"
                details_data.append(['Payment:', payment_method.upper()[:8]])
                
                # Total Amount
                details_data.append(['Total:', f"<b>₱{order.total_amount:,.2f}</b>"])
                
                # Delivery Date
                if delivery and hasattr(delivery, 'delivered_at') and delivery.delivered_at:
                    details_data.append(['Delivered:', delivery.delivered_at.strftime('%m/%d %H:%M')])
                
                # Create details table
                details_table = Table(details_data, colWidths=[35, 165])
                details_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 6),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6C757D')),
                    ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 1),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                ]))
                elements.append(details_table)
                
                # Order Items (compact)
                if hasattr(order, 'items') and order.items.exists():
                    items_text = ""
                    item_count = min(2, order.items.count())
                    for j, item in enumerate(order.items.all()[:item_count]):
                        item_name = item.item.name if hasattr(item.item, 'name') else "Item"
                        quantity = item.quantity if hasattr(item, 'quantity') else 0
                        items_text += f"{item_name[:18]}{'..' if len(item_name) > 18 else ''}×{quantity}"
                        if j < item_count - 1:
                            items_text += ", "
                    
                    if order.items.count() > 2:
                        items_text += f" +{order.items.count() - 2} more"
                    
                    items_display = Paragraph(f"<font size=5>Items: {items_text}</font>", item_style)
                    elements.append(items_display)
                
                # Notes (if any, very compact)
                if delivery and hasattr(delivery, 'notes') and delivery.notes and len(delivery.notes.strip()) > 0:
                    notes = delivery.notes.strip()
                    if len(notes) > 30:
                        notes = notes[:27] + "..."
                    notes_display = Paragraph(f"<font size=5>Note: {notes}</font>", item_style)
                    elements.append(notes_display)
                
                # Add minimal separator
                if i < len(orders) - 1:
                    elements.append(Spacer(1, 3))
                    elements.append(Paragraph("─" * 45, 
                                            ParagraphStyle('Separator', parent=styles['Normal'], 
                                                         fontSize=5, alignment=TA_CENTER, textColor=colors.lightgrey)))
                    elements.append(Spacer(1, 3))
                
            except Exception as e:
                print(f"Error processing order {order.order_id}: {str(e)}")
                continue
        
        # Add final summary
        elements.append(Spacer(1, 5))
        end_mark = Paragraph("─ END OF REPORT ─", 
                            ParagraphStyle('EndMark', parent=styles['Normal'], 
                                         fontSize=7, alignment=TA_CENTER, textColor=colors.gray))
        elements.append(end_mark)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        status_text_for_filename = status if status else 'all'
        filename = f"{status_text_for_filename}_orders_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        import traceback
        print(f"PDF generation error: {str(e)}")
        print(traceback.format_exc())
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


def generate_individual_receipts_pdf(orders):
    """Generate individual receipts PDF (compact one per page)"""
    try:
        if orders.count() == 0:
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            title = Paragraph("No Orders Found", 
                             ParagraphStyle('Title', parent=styles['Heading1'], 
                                          fontSize=14, alignment=TA_CENTER, textColor=colors.red))
            elements.append(title)
            
            subtitle = Paragraph("No orders to generate receipts.", 
                               ParagraphStyle('Subtitle', parent=styles['Normal'], 
                                            fontSize=9, alignment=TA_CENTER, textColor=colors.gray))
            elements.append(subtitle)
            
            doc.build(elements)
            buffer.seek(0)
            filename = f"receipts_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        
        buffer = BytesIO()
        
        # Create PDF document with compact margins
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=letter,
            leftMargin=0.4*inch,
            rightMargin=0.4*inch,
            topMargin=0.3*inch,
            bottomMargin=0.3*inch
        )
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        
        # Compact Receipt title style
        receipt_title_style = ParagraphStyle(
            'CompactReceiptTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#D62828'),
            alignment=TA_CENTER,
            spaceAfter=4
        )
        
        # Compact Store info style
        store_style = ParagraphStyle(
            'CompactStoreInfo',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.gray,
            spaceAfter=10
        )
        
        # Compact Order header style
        order_header_style = ParagraphStyle(
            'CompactOrderHeader',
            parent=styles['Heading2'],
            fontSize=11,
            textColor=colors.HexColor('#D62828'),
            spaceAfter=6
        )
        
        # Compact Label style
        label_style = ParagraphStyle(
            'CompactReceiptLabel',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#6C757D'),
            spaceAfter=1
        )
        
        # Compact Value style
        value_style = ParagraphStyle(
            'CompactReceiptValue',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            spaceAfter=4
        )
        
        # Process each order as individual receipt
        for order_index, order in enumerate(orders):
            try:
                delivery = order.deliveries_set.first() if hasattr(order, 'deliveries_set') and order.deliveries_set.exists() else None
                
                # Add page break except for first receipt
                if order_index > 0:
                    elements.append(PageBreak())
                
                # Receipt Header
                receipt_title = Paragraph("KUSINAEXPRESS", receipt_title_style)
                elements.append(receipt_title)
                
                store_info = Paragraph("RECEIPT", store_style)
                elements.append(store_info)
                
                elements.append(Spacer(1, 5))
                
                # Order Info
                order_header = Paragraph(f"Order #{order.order_id}", order_header_style)
                elements.append(order_header)
                
                # Status
                if delivery and hasattr(delivery, 'delivered_at') and delivery.delivered_at:
                    status = Paragraph("<b><font color='green'>✓ DELIVERED</font></b>", value_style)
                elif delivery and hasattr(delivery, 'out_for_delivery_at') and delivery.out_for_delivery_at:
                    status = Paragraph("<b><font color='blue'>OUT FOR DELIVERY</font></b>", value_style)
                elif delivery and hasattr(delivery, 'preparing_at') and delivery.preparing_at:
                    status = Paragraph("<b><font color='purple'>PREPARING</font></b>", value_style)
                elif delivery and hasattr(delivery, 'confirmed_at') and delivery.confirmed_at:
                    status = Paragraph("<b><font color='blue'>CONFIRMED</font></b>", value_style)
                else:
                    status = Paragraph("<b><font color='orange'>PENDING</font></b>", value_style)
                
                elements.append(status)
                
                elements.append(Spacer(1, 8))
                
                # Create compact details table
                details_data = []
                
                # Customer Info
                customer_name = f"{order.user.first_name or ''} {order.user.last_name or ''}".strip()
                details_data.append(['Customer:', customer_name or 'Guest'])
                
                # Order Date
                order_date = order.order_date.strftime('%b %d, %Y %I:%M %p') if hasattr(order, 'order_date') else "N/A"
                details_data.append(['Date:', order_date])
                
                # Delivery Address
                if delivery and hasattr(delivery, 'delivery_address'):
                    details_data.append(['Address:', delivery.delivery_address or '-'])
                
                # Contact Number
                if delivery and hasattr(delivery, 'contact_number'):
                    details_data.append(['Contact:', delivery.contact_number or '-'])
                
                # Create details table
                details_table = Table(details_data, colWidths=[50, 150])
                details_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6C757D')),
                    ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ]))
                elements.append(details_table)
                
                elements.append(Spacer(1, 10))
                
                # Order Items Table (compact)
                items_data = [['Item', 'Qty', 'Price', 'Total']]
                total_amount = 0
                
                if hasattr(order, 'items') and order.items.exists():
                    for item in order.items.all():
                        item_name = item.item.name if hasattr(item.item, 'name') else "Item"
                        quantity = item.quantity if hasattr(item, 'quantity') else 0
                        price = item.item.price if hasattr(item.item, 'price') else 0
                        item_total = price * quantity
                        total_amount += item_total
                        
                        # Truncate long item names
                        if len(item_name) > 25:
                            item_name = item_name[:22] + "..."
                        
                        items_data.append([
                            item_name,
                            f"×{quantity}",
                            f"₱{price:,.2f}",
                            f"₱{item_total:,.2f}"
                        ])
                else:
                    items_data.append(['No items', '', '', ''])
                
                # Create table
                table = Table(items_data, colWidths=[150, 30, 60, 60])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D62828')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
                ]))
                
                elements.append(table)
                elements.append(Spacer(1, 10))
                
                # Total Amount (compact)
                total_text = Paragraph(f"<b>TOTAL: ₱{total_amount:,.2f}</b>", 
                                      ParagraphStyle('TotalStyle', parent=styles['Normal'], 
                                                   fontSize=12, textColor=colors.HexColor('#D62828'),
                                                   alignment=TA_RIGHT))
                elements.append(total_text)
                
                # Payment Method
                payment_method = "COD"
                if hasattr(order, 'payments_set') and order.payments_set.exists():
                    payment = order.payments_set.first()
                    if payment and hasattr(payment, 'payment_method'):
                        payment_method = payment.payment_method or "COD"
                
                payment_info = Paragraph(f"<b>Payment:</b> {payment_method.upper()}", value_style)
                elements.append(payment_info)
                
                # Delivered Date
                if delivery and hasattr(delivery, 'delivered_at') and delivery.delivered_at:
                    delivered_info = Paragraph(f"<b>Delivered:</b> {delivery.delivered_at.strftime('%b %d, %Y %I:%M %p')}", value_style)
                    elements.append(delivered_info)
                
                # Notes (if any)
                if delivery and hasattr(delivery, 'notes') and delivery.notes:
                    notes = delivery.notes
                    if len(notes) > 60:
                        notes = notes[:57] + "..."
                    notes_info = Paragraph(f"<b>Note:</b> {notes}", 
                                          ParagraphStyle('Notes', parent=styles['Normal'], fontSize=8))
                    elements.append(notes_info)
                
                elements.append(Spacer(1, 15))
                
                # Compact Footer
                footer = Paragraph("Thank you for ordering!<br/>Contact: (02) 1234-5678<br/>Email: support@kusinaexpress.com", 
                                  ParagraphStyle('CompactFooter', parent=styles['Normal'], 
                                               fontSize=7, alignment=TA_CENTER, textColor=colors.gray))
                elements.append(footer)
                
            except Exception as e:
                print(f"Error processing order {order.order_id} for receipt: {str(e)}")
                continue
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"receipts_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        import traceback
        print(f"Individual receipts PDF error: {str(e)}")
        print(traceback.format_exc())
        return HttpResponse(f"Error generating individual receipts PDF: {str(e)}", status=500)


def generate_detailed_excel_report(orders, report_type, status):
    """Generate detailed Excel report with compact layout"""
    try:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        status_text = status if status else 'all'
        filename = f"{status_text}_orders_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{status_text[:20]} Orders"
        
        # Define compact styles
        header_font = Font(bold=True, size=10, color='FFFFFF')
        header_fill = PatternFill(start_color='D62828', end_color='D62828', fill_type='solid')
        title_font = Font(bold=True, size=11, color='D62828')
        compact_font = Font(size=9)
        bold_font = Font(bold=True, size=9)
        amount_font = Font(bold=True, size=9, color='D62828')
        
        # Thin border for compact cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Company Header (compact)
        ws.merge_cells('A1:H1')
        ws['A1'] = f"KUSINAEXPRESS - {status_text.upper().replace('_', ' ')} ORDERS"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Report: {datetime.now().strftime('%m/%d/%Y %I:%M%p')}"
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A3:H3')
        total_amount = sum(order.total_amount for order in orders)
        ws['A3'] = f"Orders: {orders.count()} | Total: ₱{total_amount:,.2f}"
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        
        row_num = 5
        
        if orders.count() == 0:
            ws.merge_cells(f'A{row_num}:H{row_num}')
            ws[f'A{row_num}'] = "No orders found."
            ws[f'A{row_num}'].alignment = Alignment(horizontal='center')
            ws[f'A{row_num}'].font = Font(bold=True, color='FF0000')
            
        # Create headers for compact table
        headers = ['Order#', 'Customer', 'Date', 'Status', 'Items', 'Address', 'Payment', 'Total']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        row_num += 1
        
        # Add each order as a compact row
        for order in orders:
            try:
                delivery = order.deliveries_set.first()
                
                # Customer name (compact)
                customer_name = f"{order.user.first_name or ''} {order.user.last_name or ''}".strip()
                if not customer_name:
                    customer_name = "Guest"
                if len(customer_name) > 20:
                    customer_name = customer_name[:18] + ".."
                
                # Order date (compact)
                order_date = order.order_date.strftime('%m/%d %H:%M') if hasattr(order, 'order_date') else "N/A"
                
                # Status
                if delivery and delivery.delivered_at:
                    status_text = "DELIVERED"
                elif delivery and delivery.out_for_delivery_at:
                    status_text = "OUT FOR DELIVERY"
                elif delivery and delivery.preparing_at:
                    status_text = "PREPARING"
                elif delivery and delivery.confirmed_at:
                    status_text = "CONFIRMED"
                else:
                    status_text = "PENDING"
                
                # Items count
                if hasattr(order, 'items') and order.items.exists():
                    items_count = order.items.count()
                    items_text = f"{items_count} item{'s' if items_count != 1 else ''}"
                else:
                    items_text = "0 items"
                
                # Delivery Address (truncated)
                address = "-"
                if delivery and delivery.delivery_address:
                    address = delivery.delivery_address
                    if len(address) > 25:
                        address = address[:23] + ".."
                
                # Payment Method
                payment_method = "COD"
                if hasattr(order, 'payments_set') and order.payments_set.exists():
                    payment = order.payments_set.first()
                    if payment and hasattr(payment, 'payment_method'):
                        payment_method = payment.payment_method or "COD"
                payment_text = payment_method.upper()[:10]
                
                # Create row data
                row_data = [
                    f"#{order.order_id}",
                    customer_name,
                    order_date,
                    status_text,
                    items_text,
                    address,
                    payment_text,
                    f"₱{order.total_amount:,.2f}"
                ]
                
                # Write row
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = compact_font
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                    
                    # Right align total column
                    if col_num == 8:
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.font = amount_font
                    # Center align status and items columns
                    elif col_num in [4, 5, 7]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                row_num += 1
                
                # Add order details in next row (optional, comment out for ultra-compact)
                notes = ""
                if delivery and delivery.notes and delivery.notes.strip():
                    notes = delivery.notes.strip()
                    if len(notes) > 40:
                        notes = notes[:38] + ".."
                
                if notes:
                    ws.merge_cells(f'A{row_num}:H{row_num}')
                    note_cell = ws.cell(row=row_num, column=1, value=f"Note: {notes}")
                    note_cell.font = Font(size=8, italic=True, color='666666')
                    note_cell.alignment = Alignment(horizontal='left')
                    row_num += 1
                
                # Add minimal spacing
                ws.row_dimensions[row_num].height = 5
                row_num += 1
                
            except Exception as e:
                print(f"Error processing order {order.order_id} for Excel: {str(e)}")
                continue
        
        # Auto-adjust column widths for compactness
        column_widths = [8, 20, 12, 12, 8, 25, 10, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A6'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        print(f"Excel generation error: {str(e)}")
        print(traceback.format_exc())
        return HttpResponse(f"Error generating Excel: {str(e)}", status=500)


def generate_detailed_csv_report(orders, report_type, status):
    """Generate detailed CSV report with compact format"""
    try:
        response = HttpResponse(content_type='text/csv')
        status_text = status if status else 'all'
        filename = f"{status_text}_orders_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Write compact header
        status_text = status.upper().replace('_', ' ') if status else 'ALL'
        writer.writerow([f"KUSINAEXPRESS - {status_text} ORDERS"])
        writer.writerow([f"Generated: {datetime.now().strftime('%m/%d/%Y %I:%M%p')}"])
        
        total_amount = sum(order.total_amount for order in orders)
        writer.writerow([f"Orders: {orders.count()}, Total: ₱{total_amount:,.2f}"])
        writer.writerow([])
        
        if orders.count() == 0:
            writer.writerow(["No orders found."])
            return response
        
        # Write compact column headers
        writer.writerow(['Order#', 'Customer', 'Date', 'Status', 'Items', 'Address', 'Payment', 'Total', 'Notes'])
        
        for order in orders:
            try:
                delivery = order.deliveries_set.first()
                
                # Customer name (compact)
                customer_name = f"{order.user.first_name or ''} {order.user.last_name or ''}".strip()
                if not customer_name:
                    customer_name = "Guest"
                
                # Order date (compact)
                order_date = order.order_date.strftime('%m/%d %H:%M') if hasattr(order, 'order_date') else "N/A"
                
                # Status
                if delivery and delivery.delivered_at:
                    status_text = "DELIVERED"
                elif delivery and delivery.out_for_delivery_at:
                    status_text = "OUT FOR DELIVERY"
                elif delivery and delivery.preparing_at:
                    status_text = "PREPARING"
                elif delivery and delivery.confirmed_at:
                    status_text = "CONFIRMED"
                else:
                    status_text = "PENDING"
                
                # Items count
                if hasattr(order, 'items') and order.items.exists():
                    items_count = order.items.count()
                    items_text = f"{items_count} item{'s' if items_count != 1 else ''}"
                else:
                    items_text = "0 items"
                
                # Delivery Address (truncated)
                address = "-"
                if delivery and delivery.delivery_address:
                    address = delivery.delivery_address
                    if len(address) > 30:
                        address = address[:28] + ".."
                
                # Payment Method
                payment_method = "COD"
                if hasattr(order, 'payments_set') and order.payments_set.exists():
                    payment = order.payments_set.first()
                    if payment and hasattr(payment, 'payment_method'):
                        payment_method = payment.payment_method or "COD"
                payment_text = payment_method.upper()
                
                # Notes (truncated)
                notes = ""
                if delivery and delivery.notes and delivery.notes.strip():
                    notes = delivery.notes.strip()
                    if len(notes) > 30:
                        notes = notes[:28] + ".."
                
                # Create compact row
                row = [
                    f"#{order.order_id}",
                    customer_name,
                    order_date,
                    status_text,
                    items_text,
                    address,
                    payment_text,
                    f"₱{order.total_amount:,.2f}",
                    notes
                ]
                
                writer.writerow(row)
                
            except Exception as e:
                print(f"Error processing order {order.order_id} for CSV: {str(e)}")
                continue
        
        # Add summary row
        writer.writerow([])
        writer.writerow(['SUMMARY', f'Total Orders: {orders.count()}', f'Total Amount: ₱{total_amount:,.2f}'])
        
        return response
        
    except Exception as e:
        import traceback
        print(f"CSV generation error: {str(e)}")
        print(traceback.format_exc())
        return HttpResponse(f"Error generating CSV: {str(e)}", status=500)